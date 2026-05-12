"""Merge two CSV/TSV files by key columns, selecting columns from each side.

Typical use is to keep all columns from a base file and add one or more columns
from a second file by joining on key column(s).

Usage::

    python -m src.merge_columns base.tsv extra.tsv out.tsv --key-col id --add-col score
    python -m src.merge_columns base.tsv extra.tsv out.tsv --key-col gene,variant --add-all-cols-from-extra

    # Rename a column while adding it:
    python -m src.merge_columns base.tsv metadata.xlsx out.tsv \\
        --key-col dataset_name \\
        --extra-worksheet Curation \\
        --add-col "Detects Splicing Variants?:splice_measure" \\
        --add-col "Ensembl Transcript ID"

    # Key column has different names in the two files (base:extra syntax):
    python -m src.merge_columns base.tsv meta.xlsx out.tsv \\
        --key-col "dataset_name:Dataset Name" \\
        --add-col "Ensembl Transcript ID"
"""

from __future__ import annotations

import csv
from pathlib import Path

import click


def _detect_separator(file_path: str) -> str:
    return "\t" if Path(file_path).suffix.lower() in (".tsv", ".txt") else ","


def _split_csv_args(values: tuple[str, ...]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in values:
        for item in raw.split(","):
            col = item.strip()
            if not col:
                continue
            if col not in seen:
                seen.add(col)
                out.append(col)
    return out


def _parse_add_col_args(
    values: tuple[str, ...],
) -> tuple[list[str], dict[str, str]]:
    """Parse ``--add-col`` values that may include ``SRC:DEST`` rename syntax.

    Returns ``(add_columns, column_renames)`` where *add_columns* is a
    deduplicated list of source column names and *column_renames* maps each
    source name to its desired output name (only entries that differ).

    Each value may be comma-separated (``"A,B:b_out,C"``) and/or repeated via
    multiple ``--add-col`` flags.  A bare colon is not a valid separator inside
    a column name, so ``SRC:DEST`` is unambiguous.
    """
    add_columns: list[str] = []
    column_renames: dict[str, str] = {}
    seen: set[str] = set()
    for raw in values:
        for item in raw.split(","):
            item = item.strip()
            if not item:
                continue
            if ":" in item:
                src, _, dest = item.partition(":")
                src = src.strip()
                dest = dest.strip()
            else:
                src = item
                dest = item
            if src not in seen:
                seen.add(src)
                add_columns.append(src)
                if dest != src:
                    column_renames[src] = dest
    return add_columns, column_renames


def _parse_key_col_args(
    values: tuple[str, ...],
) -> tuple[list[str], list[str]]:
    """Parse ``--key-col`` values that may include ``BASE_COL:EXTRA_COL`` rename syntax.

    Returns ``(base_key_columns, extra_key_columns)``.  When a value has no
    colon the same name is used on both sides, so the two lists are always the
    same length.  Each value may also be comma-separated.
    """
    base_cols: list[str] = []
    extra_cols: list[str] = []
    seen: set[str] = set()
    for raw in values:
        for item in raw.split(","):
            item = item.strip()
            if not item:
                continue
            if ":" in item:
                base, _, extra = item.partition(":")
                base = base.strip()
                extra = extra.strip()
            else:
                base = item
                extra = item
            if base not in seen:
                seen.add(base)
                base_cols.append(base)
                extra_cols.append(extra)
    return base_cols, extra_cols


def _build_key(row: dict[str, str], key_columns: list[str]) -> tuple[str, ...]:
    return tuple((row.get(col) or "").strip() for col in key_columns)


def _read_excel_extra(
    file_path: str,
    worksheet: str | None,
) -> tuple[list[str], list[dict[str, str]]]:
    """Read an Excel workbook and return ``(fieldnames, rows)`` as string dicts.

    *worksheet* may be a sheet name or a 1-based integer index string.
    When *worksheet* is ``None`` the workbook's active (first) sheet is used.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read Excel files. "
            "Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if worksheet is None:
            ws = wb.active
        elif worksheet in wb.sheetnames:
            ws = wb[worksheet]
        else:
            try:
                idx = int(worksheet) - 1
                ws = wb.worksheets[idx]
            except (ValueError, IndexError):
                raise ValueError(
                    f"Worksheet '{worksheet}' not found in {file_path!r}. "
                    f"Available sheets: {wb.sheetnames}"
                )

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return [], []

        fieldnames = [str(h).strip() if h is not None else "" for h in header]

        rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            row_dict: dict[str, str] = {}
            for col, val in zip(fieldnames, raw_row):
                row_dict[col] = "" if val is None else str(val)
            rows.append(row_dict)
    finally:
        wb.close()

    return fieldnames, rows


def _is_excel(file_path: str) -> bool:
    return Path(file_path).suffix.lower() in (".xlsx", ".xls", ".xlsm")


def merge_columns(
    base_file: str,
    extra_file: str,
    output_file: str,
    key_columns: list[str],
    add_columns: list[str],
    add_all_cols_from_extra: bool = False,
    column_renames: dict[str, str] | None = None,
    extra_worksheet: str | None = None,
    extra_key_columns: list[str] | None = None,
) -> int:
    """Left-join *base_file* with *extra_file* and write selected merged columns.

    The output always contains all base-file rows. The join uses *key_columns*
    (from the base file).  When the same join column has a different name in
    *extra_file*, pass *extra_key_columns* with the corresponding names.  If
    *extra_key_columns* is ``None`` or empty, *key_columns* is used for both
    sides (original behaviour).

    *extra_file* may be a CSV/TSV file or an Excel workbook (.xlsx/.xls/.xlsm).
    When *extra_file* is an Excel workbook, *extra_worksheet* selects the sheet
    (by name or 1-based integer index; defaults to the first/active sheet).

    *column_renames* maps source column names to their desired output names.
    This can also be expressed inline in *add_columns* as ``"SRC:DEST"`` strings
    when using the CLI (parsed by :func:`_parse_add_col_args`).
    """
    if not key_columns:
        raise ValueError("At least one key column is required.")

    extra_key_cols: list[str] = extra_key_columns if extra_key_columns else key_columns
    if len(extra_key_cols) != len(key_columns):
        raise ValueError(
            f"extra_key_columns length ({len(extra_key_cols)}) must match "
            f"key_columns length ({len(key_columns)})."
        )

    renames: dict[str, str] = column_renames or {}

    base_sep = _detect_separator(base_file)

    with open(base_file, newline="", encoding="utf-8") as base_fh:
        base_reader = csv.DictReader(base_fh, delimiter=base_sep)
        base_fieldnames = list(base_reader.fieldnames or [])
        if not base_fieldnames:
            raise ValueError(f"Input file {base_file!r} is empty or missing a header.")
        missing_base_keys = [c for c in key_columns if c not in base_fieldnames]
        if missing_base_keys:
            raise ValueError(f"Base file is missing key columns: {missing_base_keys}")
        base_rows = list(base_reader)

    if _is_excel(extra_file):
        extra_fieldnames, extra_rows = _read_excel_extra(extra_file, extra_worksheet)
    else:
        extra_sep = _detect_separator(extra_file)
        with open(extra_file, newline="", encoding="utf-8") as extra_fh:
            extra_reader = csv.DictReader(extra_fh, delimiter=extra_sep)
            extra_fieldnames = list(extra_reader.fieldnames or [])
            extra_rows = list(extra_reader)

    if not extra_fieldnames:
        raise ValueError(f"Input file {extra_file!r} is empty or missing a header.")
    missing_extra_keys = [c for c in extra_key_cols if c not in extra_fieldnames]
    if missing_extra_keys:
        raise ValueError(f"Extra file is missing key columns: {missing_extra_keys}")

    if add_all_cols_from_extra:
        selected_add_columns = [
            c for c in extra_fieldnames
            if c not in extra_key_cols and renames.get(c, c) not in base_fieldnames
        ]
    elif not add_columns:
        selected_add_columns = [
            c for c in extra_fieldnames
            if c not in extra_key_cols and renames.get(c, c) not in base_fieldnames
        ]
    else:
        missing_add_cols = [c for c in add_columns if c not in extra_fieldnames]
        if missing_add_cols:
            raise ValueError(f"Extra file is missing add columns: {missing_add_cols}")
        selected_add_columns = add_columns

    extra_by_key: dict[tuple[str, ...], dict[str, str]] = {}
    for row in extra_rows:
        extra_by_key[_build_key(row, extra_key_cols)] = row

    # Build output fieldnames, applying any renames
    output_fieldnames = list(base_fieldnames)
    for src_col in selected_add_columns:
        dest_col = renames.get(src_col, src_col)
        if dest_col not in output_fieldnames:
            output_fieldnames.append(dest_col)

    out_sep = _detect_separator(output_file)
    with open(output_file, "w", newline="", encoding="utf-8") as out_fh:
        writer = csv.DictWriter(
            out_fh,
            fieldnames=output_fieldnames,
            delimiter=out_sep,
            extrasaction="ignore",
        )
        writer.writeheader()

        for base_row in base_rows:
            merged = dict(base_row)
            extra_row = extra_by_key.get(_build_key(base_row, key_columns))
            for src_col in selected_add_columns:
                dest_col = renames.get(src_col, src_col)
                merged[dest_col] = extra_row.get(src_col, "") if extra_row is not None else ""
            writer.writerow(merged)

    return len(base_rows)


@click.command()
@click.argument("base_file")
@click.argument("extra_file")
@click.argument("output_file")
@click.option(
    "--key-col",
    "key_cols_raw",
    multiple=True,
    required=True,
    metavar="BASE_COL[:EXTRA_COL]",
    help=(
        "Key column(s) for join identity. May be repeated and/or comma-separated. "
        "Use BASE_COL:EXTRA_COL when the column has different names in the two files "
        "(e.g. --key-col 'dataset_name:Dataset Name'). "
        "Plain names mean the same column is used on both sides."
    ),
)
@click.option(
    "--add-col",
    "add_cols_raw",
    multiple=True,
    metavar="COLUMN[:OUTPUT_NAME]",
    help=(
        "Column(s) to add from extra file. May be repeated and/or comma-separated. "
        "Use SRC:DEST to rename a column in the output "
        "(e.g. --add-col 'Detects Splicing Variants?:splice_measure'). "
        "If omitted, all new non-key columns from extra are added."
    ),
)
@click.option(
    "--add-all-cols-from-extra",
    is_flag=True,
    help="Add all non-key columns from extra that are not already in base.",
)
@click.option(
    "--extra-worksheet",
    default=None,
    metavar="NAME_OR_INDEX",
    help=(
        "Worksheet to read when EXTRA_FILE is an Excel workbook (.xlsx/.xls). "
        "Accepts a sheet name or a 1-based integer index. "
        "Defaults to the first (active) sheet."
    ),
)
@click.option(
    "--csv-field-size-limit",
    default=csv.field_size_limit(),
    show_default=True,
    type=click.IntRange(min=1),
    help="Maximum per-field character length for CSV/TSV parsing.",
)
def main(
    base_file: str,
    extra_file: str,
    output_file: str,
    key_cols_raw: tuple[str, ...],
    add_cols_raw: tuple[str, ...],
    add_all_cols_from_extra: bool,
    extra_worksheet: str | None,
    csv_field_size_limit: int,
) -> None:
    csv.field_size_limit(csv_field_size_limit)
    key_columns, extra_key_columns = _parse_key_col_args(key_cols_raw)
    add_columns, column_renames = _parse_add_col_args(add_cols_raw)

    try:
        n_rows = merge_columns(
            base_file=base_file,
            extra_file=extra_file,
            output_file=output_file,
            key_columns=key_columns,
            add_columns=add_columns,
            add_all_cols_from_extra=add_all_cols_from_extra,
            column_renames=column_renames,
            extra_worksheet=extra_worksheet,
            extra_key_columns=extra_key_columns,
        )
    except (ValueError, ImportError) as exc:
        raise click.ClickException(str(exc)) from exc

    click.echo(f"Wrote {n_rows} row(s) to {output_file}")


if __name__ == "__main__":
    main()
