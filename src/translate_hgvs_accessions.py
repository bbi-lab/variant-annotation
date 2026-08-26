"""Translate accessions embedded in HGVS columns between RefSeq and Ensembl namespaces.

Handles three kinds of columns independently, each opted into via a CLI flag:

  --hgvs-c-col COL
      Column holding qualified coding HGVS (``ACCESSION:c.change``). Two modes:
        --hgvs-c-mode translate (default)  Swap the transcript accession for its
                                            MANE-equivalent in the opposite namespace
                                            (direction controlled by --direction).
                                            Errors out if the accession is not a
                                            MANE Select/Plus Clinical transcript.
        --hgvs-c-mode keep                 Leave the accession as-is (only
                                            --normalize-version-separator applies;
                                            no --mane-file/--direction needed).

  --hgvs-g-col COL
      Column holding qualified genomic HGVS (``ACCESSION:g.change``). Two modes:
        --hgvs-g-mode keep        Leave the accession as-is (only version-separator
                                   normalization, if requested, is applied).
        --hgvs-g-mode chrom-name  Replace the accession with the bare Ensembl-style
                                   chromosome name (e.g. "7", "X", "MT"), via an
                                   embedded RefSeq-accession -> chromosome-name table.

  --hgvs-p-col COL
      Column holding protein HGVS. Two modes:
        --hgvs-p-mode translate  Cell already has an accession prefix
                                  (``ACCESSION:p.change``); swap it via the MANE
                                  mapping, like --hgvs-c-col.
        --hgvs-p-mode add        Cell is an unqualified change (``p.change``); a
                                  protein accession is inferred from the transcript
                                  accession in --hgvs-c-col (same row) and prepended.
                                  --hgvs-p-accession chooses refseq or ensembl.
                                  Cells that don't start with "p." (e.g. "splice
                                  region" placeholders for intronic variants) are
                                  passed through unchanged.

Some source data encodes the HGVS version suffix with an underscore instead of a
dot (e.g. ``NM_032415_7`` for ``NM_032415.7``). Pass --normalize-version-separator
to accept either form on input; output always uses the standard dot form.

All other columns, and column order, are passed through unchanged. Input/output
format (csv, tsv, xlsx) is inferred from the file extension unless overridden with
--input-format/--output-format.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# RefSeq chromosome accession -> Ensembl-style chromosome name (GRCh38)
# ---------------------------------------------------------------------------

CHROM_BY_ACCESSION: dict[str, str] = {
    "NC_000001.11": "1",
    "NC_000002.12": "2",
    "NC_000003.12": "3",
    "NC_000004.12": "4",
    "NC_000005.10": "5",
    "NC_000006.12": "6",
    "NC_000007.14": "7",
    "NC_000008.11": "8",
    "NC_000009.12": "9",
    "NC_000010.11": "10",
    "NC_000011.10": "11",
    "NC_000012.12": "12",
    "NC_000013.11": "13",
    "NC_000014.9": "14",
    "NC_000015.10": "15",
    "NC_000016.10": "16",
    "NC_000017.11": "17",
    "NC_000018.12": "18",
    "NC_000019.10": "19",
    "NC_000020.11": "20",
    "NC_000021.9": "21",
    "NC_000022.11": "22",
    "NC_000023.11": "X",
    "NC_000024.10": "Y",
    "NC_012920.1": "MT",
}

CHROM_BY_BASE_ACCESSION: dict[str, str] = {
    acc.rsplit(".", 1)[0]: name for acc, name in CHROM_BY_ACCESSION.items()
}


# ---------------------------------------------------------------------------
# Version-separator normalization
# ---------------------------------------------------------------------------

# Matches accessions like "NM_032415_7" or "NM_032415.7" -> ("NM_032415", "7").
_VERSIONED_ACCESSION_RE = re.compile(r"^([A-Za-z]+_[A-Za-z0-9]+)[._](\d+)$")


def _normalize_accession(accession: str) -> str:
    """Rewrite an underscore-versioned accession (``NM_032415_7``) to dot form.

    Accessions already in dot form, or without a recognizable version suffix,
    are returned unchanged.
    """
    match = _VERSIONED_ACCESSION_RE.match(accession)
    if match:
        return f"{match.group(1)}.{match.group(2)}"
    return accession


# ---------------------------------------------------------------------------
# Malformed multi-variant (haplotype) syntax normalization
# ---------------------------------------------------------------------------

# A correctly-formed multi-variant HGVS description wraps the semicolon-joined
# changes in brackets: "c.[2523A>C;2533A>T]". Source data sometimes has the
# malformed, unbracketed form instead: "c.2523A>C;2533A>T" (the "c." prefix is
# only present once, on the first change).
_ALREADY_BRACKETED_RE = re.compile(r"^[cgpn]\.\[.*\]$")
_KIND_PREFIX_RE = re.compile(r"^([cgpn])\.(.+)$")


def _normalize_multi_variant(rest: str) -> str:
    """Rewrite a malformed unbracketed multi-variant description to bracket form.

    ``"c.2523A>C;2533A>T"`` -> ``"c.[2523A>C;2533A>T]"``. Already-bracketed,
    single-variant, or otherwise-unrecognized strings are returned unchanged.
    """
    if ";" not in rest or _ALREADY_BRACKETED_RE.match(rest):
        return rest
    match = _KIND_PREFIX_RE.match(rest)
    if not match:
        return rest
    kind, body = match.group(1), match.group(2)
    if "[" in body or "]" in body:
        return rest
    components = [c.strip() for c in body.split(";")]
    return f"{kind}.[{';'.join(components)}]"


# ---------------------------------------------------------------------------
# MANE mapping
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class ManeRecord:
    refseq_nuc: str
    refseq_prot: str
    ensembl_nuc: str
    ensembl_prot: str


def _open_maybe_gzipped(path: str):
    if path.endswith(".gz"):
        return gzip.open(path, "rt", encoding="utf-8")
    return open(path, newline="", encoding="utf-8")


def _build_mane_index(mane_file: Optional[str]) -> tuple[dict[str, ManeRecord], dict[str, ManeRecord]]:
    """Build accession -> ManeRecord indexes from an NCBI MANE summary file.

    Returns ``(by_accession, by_base_accession)``. *by_accession* is keyed by
    versioned RefSeq_nuc/RefSeq_prot/Ensembl_nuc/Ensembl_prot accessions.
    *by_base_accession* is a version-stripped fallback (versioned takes priority).
    """
    assert mane_file is not None
    by_accession: dict[str, ManeRecord] = {}
    with _open_maybe_gzipped(mane_file) as fh:
        reader = csv.DictReader(fh, delimiter="\t")
        for row in reader:
            refseq_nuc = (row.get("RefSeq_nuc") or "").strip()
            refseq_prot = (row.get("RefSeq_prot") or "").strip()
            ensembl_nuc = (row.get("Ensembl_nuc") or "").strip()
            ensembl_prot = (row.get("Ensembl_prot") or "").strip()
            if not refseq_nuc or not ensembl_nuc:
                continue
            record = ManeRecord(refseq_nuc, refseq_prot, ensembl_nuc, ensembl_prot)
            by_accession[refseq_nuc] = record
            by_accession[ensembl_nuc] = record

    by_base_accession: dict[str, ManeRecord] = {}
    for acc, record in by_accession.items():
        if "." in acc:
            base = acc.rsplit(".", 1)[0]
            by_base_accession.setdefault(base, record)

    logger.info("Loaded MANE mapping with %d accession keys.", len(by_accession))
    return by_accession, by_base_accession


def _lookup_mane_record(
    accession: str,
    by_accession: dict[str, ManeRecord],
    by_base_accession: dict[str, ManeRecord],
) -> Optional[ManeRecord]:
    if accession in by_accession:
        return by_accession[accession]
    if "." in accession:
        base = accession.rsplit(".", 1)[0]
        if base in by_base_accession:
            return by_base_accession[base]
    return None


# ---------------------------------------------------------------------------
# Per-cell translation
# ---------------------------------------------------------------------------

class TranslationError(RuntimeError):
    """Raised when an accession cannot be translated (e.g. not a MANE transcript)."""


def _split_hgvs(value: str) -> Optional[tuple[str, str]]:
    if not value or ":" not in value:
        return None
    accession, rest = value.split(":", 1)
    return accession, _normalize_multi_variant(rest)


def _translate_hgvs_c(
    value: str,
    mode: str,
    direction: Optional[str],
    by_accession: dict[str, ManeRecord],
    by_base_accession: dict[str, ManeRecord],
    normalize_versions: bool,
    row_context: str,
) -> str:
    split = _split_hgvs(value)
    if split is None:
        return value
    accession, rest = split
    if normalize_versions:
        accession = _normalize_accession(accession)
    if mode == "keep":
        return f"{accession}:{rest}"

    # mode == "translate"
    record = _lookup_mane_record(accession, by_accession, by_base_accession)
    if record is None:
        raise TranslationError(
            f"{row_context}: transcript accession {accession!r} not found in MANE file "
            "(not a MANE Select/Plus Clinical transcript?)"
        )
    target = record.ensembl_nuc if direction == "to-ensembl" else record.refseq_nuc
    return f"{target}:{rest}"


def _translate_hgvs_g(
    value: str,
    mode: str,
    normalize_versions: bool,
    row_context: str,
) -> str:
    split = _split_hgvs(value)
    if split is None:
        return value
    accession, rest = split
    if normalize_versions:
        accession = _normalize_accession(accession)
    if mode == "keep":
        return f"{accession}:{rest}"

    # mode == "chrom-name"
    chrom_name = CHROM_BY_ACCESSION.get(accession)
    if chrom_name is None and "." in accession:
        chrom_name = CHROM_BY_BASE_ACCESSION.get(accession.rsplit(".", 1)[0])
    if chrom_name is None:
        raise TranslationError(
            f"{row_context}: chromosome accession {accession!r} not found in the "
            "RefSeq-accession -> chromosome-name table"
        )
    return f"{chrom_name}:{rest}"


def _translate_hgvs_p(
    value: str,
    mode: Optional[str],
    direction: Optional[str],
    accession_kind: Optional[str],
    source_transcript_accession: Optional[str],
    normalize_versions: bool,
    by_accession: dict[str, ManeRecord],
    by_base_accession: dict[str, ManeRecord],
    row_context: str,
) -> str:
    if not value:
        return value

    if mode == "translate":
        split = _split_hgvs(value)
        if split is None:
            return value
        accession, rest = split
        if normalize_versions:
            accession = _normalize_accession(accession)
        record = _lookup_mane_record(accession, by_accession, by_base_accession)
        if record is None:
            raise TranslationError(
                f"{row_context}: protein accession {accession!r} not found in MANE file"
            )
        target = record.ensembl_prot if direction == "to-ensembl" else record.refseq_prot
        return f"{target}:{rest}"

    # mode == "add"
    if not value.startswith("p."):
        return value
    if source_transcript_accession is None:
        raise TranslationError(
            f"{row_context}: cannot add a protein accession; coding-variant column "
            "value is missing or unqualified"
        )
    accession = source_transcript_accession
    if normalize_versions:
        accession = _normalize_accession(accession)
    record = _lookup_mane_record(accession, by_accession, by_base_accession)
    if record is None:
        raise TranslationError(
            f"{row_context}: transcript accession {accession!r} (from coding-variant "
            "column) not found in MANE file"
        )
    protein_accession = record.ensembl_prot if accession_kind == "ensembl" else record.refseq_prot
    if not protein_accession:
        raise TranslationError(
            f"{row_context}: MANE record for {accession!r} has no {accession_kind} "
            "protein accession"
        )
    return f"{protein_accession}:{value}"


# ---------------------------------------------------------------------------
# File I/O (csv, tsv, xlsx)
# ---------------------------------------------------------------------------

_FORMAT_BY_SUFFIX = {".xlsx": "xlsx", ".xls": "xlsx", ".tsv": "tsv", ".csv": "csv"}


def _detect_format(path: str, explicit: Optional[str]) -> str:
    if explicit and explicit != "auto":
        return explicit
    suffix = Path(path).suffix.lower()
    fmt = _FORMAT_BY_SUFFIX.get(suffix)
    if fmt is None:
        raise ValueError(
            f"Cannot infer format from extension {suffix!r} for {path!r}; "
            "pass --input-format/--output-format explicitly."
        )
    return fmt


def _read_excel(path: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.active
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            try:
                idx = int(sheet) - 1
                ws = wb.worksheets[idx]
            except (ValueError, IndexError):
                raise ValueError(
                    f"Worksheet {sheet!r} not found in {path!r}. Available sheets: {wb.sheetnames}"
                )

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = [str(h).strip() if h is not None else "" for h in header]

        rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue
            rows.append({col: ("" if val is None else str(val)) for col, val in zip(fieldnames, raw_row)})
    finally:
        wb.close()

    return fieldnames, rows


def _read_delimited(path: str, sep: str) -> tuple[list[str], list[dict[str, str]]]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh, delimiter=sep)
        if reader.fieldnames is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = list(reader.fieldnames)
        rows = [dict(row) for row in reader]
    return fieldnames, rows


def _read_rows(path: str, fmt: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    if fmt == "xlsx":
        return _read_excel(path, sheet)
    return _read_delimited(path, "\t" if fmt == "tsv" else ",")


def _write_excel(path: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])
    wb.save(path)


def _write_delimited(path: str, sep: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, delimiter=sep, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _drop_blank_columns(fieldnames: list[str], rows: list[dict[str, str]]) -> list[str]:
    """Return *fieldnames* with blank columns removed.

    A column is considered blank if its header is empty/whitespace-only and it
    holds no non-empty value in any row (e.g. a stray unnamed trailing column
    left over from an Excel export).
    """
    kept: list[str] = []
    for col in fieldnames:
        if col.strip():
            kept.append(col)
            continue
        if any((row.get(col) or "").strip() for row in rows):
            kept.append(col)
            continue
        logger.info("Dropping blank column %r (empty header, no data in any row).", col)
    return kept


def _write_rows(path: str, fmt: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    if fmt == "xlsx":
        _write_excel(path, fieldnames, rows)
        return
    _write_delimited(path, "\t" if fmt == "tsv" else ",", fieldnames, rows)


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def translate_hgvs_accessions(
    input_file: str,
    output_file: str,
    *,
    input_format: Optional[str] = None,
    output_format: Optional[str] = None,
    sheet: Optional[str] = None,
    mane_file: Optional[str] = None,
    direction: Optional[str] = None,
    hgvs_c_col: Optional[str] = None,
    hgvs_c_mode: str = "translate",
    hgvs_g_col: Optional[str] = None,
    hgvs_g_mode: str = "keep",
    hgvs_p_col: Optional[str] = None,
    hgvs_p_mode: Optional[str] = None,
    hgvs_p_accession: Optional[str] = None,
    normalize_versions: bool = False,
    drop_blank_columns: bool = False,
) -> None:
    in_fmt = _detect_format(input_file, input_format)
    out_fmt = _detect_format(output_file, output_format)

    fieldnames, rows = _read_rows(input_file, in_fmt, sheet)

    if drop_blank_columns:
        fieldnames = _drop_blank_columns(fieldnames, rows)

    for col in (hgvs_c_col, hgvs_g_col, hgvs_p_col):
        if col and col not in fieldnames:
            raise ValueError(f"Column {col!r} not found in input; available columns: {fieldnames}")

    if hgvs_p_col and hgvs_p_mode == "add" and not hgvs_c_col:
        raise ValueError("--hgvs-p-mode add requires --hgvs-c-col to infer the transcript accession from.")

    translates_c = bool(hgvs_c_col) and hgvs_c_mode == "translate"
    translates_p = bool(hgvs_p_col) and hgvs_p_mode in ("translate", "add")
    needs_mane = translates_c or translates_p
    needs_direction = translates_c or (hgvs_p_col and hgvs_p_mode == "translate")

    if needs_mane and not mane_file:
        raise ValueError("--mane-file is required when translating --hgvs-c-col (mode=translate) or --hgvs-p-col.")
    if needs_direction and not direction:
        raise ValueError("--direction is required when --hgvs-c-col is in 'translate' mode, or --hgvs-p-col is in 'translate' mode.")
    if hgvs_p_col and hgvs_p_mode == "add" and not hgvs_p_accession:
        raise ValueError("--hgvs-p-accession is required when --hgvs-p-mode is 'add'.")

    by_accession: dict[str, ManeRecord] = {}
    by_base_accession: dict[str, ManeRecord] = {}
    if needs_mane:
        by_accession, by_base_accession = _build_mane_index(mane_file)

    for i, row in enumerate(rows, start=2):
        row_context = f"row {i}"

        source_transcript_accession: Optional[str] = None
        if hgvs_c_col:
            original_c = row[hgvs_c_col]
            split = _split_hgvs(original_c)
            if split is not None:
                source_transcript_accession = split[0]
            row[hgvs_c_col] = _translate_hgvs_c(
                original_c, hgvs_c_mode, direction, by_accession, by_base_accession, normalize_versions, row_context
            )

        if hgvs_g_col:
            row[hgvs_g_col] = _translate_hgvs_g(
                row[hgvs_g_col], hgvs_g_mode, normalize_versions, row_context
            )

        if hgvs_p_col:
            row[hgvs_p_col] = _translate_hgvs_p(
                row[hgvs_p_col],
                hgvs_p_mode,
                direction,
                hgvs_p_accession,
                source_transcript_accession,
                normalize_versions,
                by_accession,
                by_base_accession,
                row_context,
            )

    _write_rows(output_file, out_fmt, fieldnames, rows)
    logger.info("Wrote %d rows to %s", len(rows), output_file)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Translate accessions in hgvs_c/hgvs_g/hgvs_p columns between RefSeq and "
            "Ensembl namespaces, using an NCBI MANE summary file. Reads/writes csv, "
            "tsv, or xlsx."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("input", help="Input file (csv, tsv, or xlsx).")
    p.add_argument("output", help="Output file (csv, tsv, or xlsx).")

    p.add_argument("--input-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto",
                    help="Override input format detection (default: infer from extension).")
    p.add_argument("--output-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto",
                    help="Override output format detection (default: infer from extension).")
    p.add_argument("--sheet", default=None,
                    help="Worksheet name or 1-based index for xlsx input (default: active/first sheet).")

    p.add_argument("--mane-file", default=None, metavar="FILE",
                    help="NCBI MANE summary file (MANE.GRCh38.*.summary.txt[.gz]). "
                         "Required if --hgvs-c-col is given, or --hgvs-p-col uses translate/add mode.")
    p.add_argument("--direction", choices=["to-refseq", "to-ensembl"], default=None,
                    help="Swap direction for accession translation. Required for --hgvs-c-col and "
                         "for --hgvs-p-mode translate.")

    p.add_argument("--hgvs-c-col", default=None, metavar="COL",
                    help="Column with qualified coding HGVS (ACCESSION:c.change) to translate.")
    p.add_argument("--hgvs-c-mode", choices=["translate", "keep"], default="translate",
                    help="'translate' swaps the transcript accession via MANE/--direction; "
                         "'keep' leaves it as-is (only --normalize-version-separator applies, "
                         "no --mane-file/--direction needed).")

    p.add_argument("--hgvs-g-col", default=None, metavar="COL",
                    help="Column with qualified genomic HGVS (ACCESSION:g.change).")
    p.add_argument("--hgvs-g-mode", choices=["keep", "chrom-name"], default="keep",
                    help="'keep' leaves the genomic accession as-is (RefSeq NC_ accession); "
                         "'chrom-name' replaces it with the bare Ensembl-style chromosome name.")

    p.add_argument("--hgvs-p-col", default=None, metavar="COL",
                    help="Column with protein HGVS (qualified or bare p.change).")
    p.add_argument("--hgvs-p-mode", choices=["translate", "add"], default=None,
                    help="'translate' swaps an existing accession prefix; 'add' prepends a protein "
                         "accession inferred from --hgvs-c-col (bare p.change cells only; other "
                         "values are passed through unchanged).")
    p.add_argument("--hgvs-p-accession", choices=["refseq", "ensembl"], default=None,
                    help="Which protein accession flavor to prepend when --hgvs-p-mode is 'add'.")

    p.add_argument("--normalize-version-separator", action="store_true", dest="normalize_versions",
                    help="Accept '_' as well as '.' before the version number in input accessions "
                         "(e.g. NM_032415_7). Output always uses '.'.")
    p.add_argument("--drop-blank-columns", action="store_true",
                    help="Drop columns whose header is blank and that hold no data in any row "
                         "(e.g. a stray unnamed trailing column from an Excel export).")

    p.add_argument("--csv-field-size-limit", type=int, default=131072,
                    help="Maximum field size for csv/tsv parsing.")
    p.add_argument("--log-level", default="WARNING",
                    choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
                    help="Logging verbosity.")
    return p


def main(argv: Optional[list[str]] = None) -> None:
    parser = _build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        stream=sys.stderr,
    )

    csv.field_size_limit(args.csv_field_size_limit)

    translate_hgvs_accessions(
        args.input,
        args.output,
        input_format=args.input_format,
        output_format=args.output_format,
        sheet=args.sheet,
        mane_file=args.mane_file,
        direction=args.direction,
        hgvs_c_col=args.hgvs_c_col,
        hgvs_c_mode=args.hgvs_c_mode,
        hgvs_g_col=args.hgvs_g_col,
        hgvs_g_mode=args.hgvs_g_mode,
        hgvs_p_col=args.hgvs_p_col,
        hgvs_p_mode=args.hgvs_p_mode,
        hgvs_p_accession=args.hgvs_p_accession,
        normalize_versions=args.normalize_versions,
        drop_blank_columns=args.drop_blank_columns,
    )


if __name__ == "__main__":
    main()
