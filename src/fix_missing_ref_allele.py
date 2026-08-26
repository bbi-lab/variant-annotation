"""Insert a missing reference-allele letter in malformed coding-HGVS substitutions.

Some source rows are missing the reference letter in a single-nucleotide
substitution call, e.g. ``NM_032415_7:c.865-4>G`` instead of the well-formed
``NM_032415_7:c.865-4C>G`` — a manual transcription typo (compare sibling rows
at the same position, which do have the reference letter: ``c.865-4C>A``,
``c.865-4C>T``).

Since the row's own ``seq``/``reference`` columns record the observed vs.
assay-reference base at every position in the local window, the missing
letter can be reconstructed directly: find the single mismatch position
between ``seq`` and ``reference`` for that row, and take
``reference[offset]`` as the missing reference letter. The malformed cell's
own alternate letter is cross-checked against ``seq[offset]`` as a sanity
check before accepting the reconstruction.

Only cells matching this exact malformed shape (a coding, UTR, or
intronic-offset position with an alternate but no reference letter) are
touched. Other cells (haplotypes, indels, "c.=", well-formed substitutions,
etc.) are passed through unchanged.

Note the reconstructed reference letter comes from the assay's local
*reference* column, not necessarily the true transcript sequence — a
downstream fixed-edit reference check (``correct_reference_alleles.py``)
still applies to the repaired cell like any other.
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# Matches a substitution missing its reference letter: c.<pos>[+-offset]>ALT
_MISSING_REF_RE = re.compile(
    r"^c\.(?P<utr>[-*])?(?P<base>\d+)(?:(?P<isign>[-+])(?P<offset>\d+))?>(?P<alt>[ACGTNacgtn])$"
)


def _mismatch_offset(seq: str, ref: str) -> Optional[int]:
    """Return the single 0-based index where *seq* differs from *ref*, or None."""
    if len(seq) != len(ref):
        return None
    offsets = [i for i, (a, b) in enumerate(zip(seq, ref)) if a != b]
    return offsets[0] if len(offsets) == 1 else None


# ---------------------------------------------------------------------------
# File I/O (csv, tsv, xlsx) — same conventions as translate_hgvs_accessions.py
# ---------------------------------------------------------------------------

_FORMAT_BY_SUFFIX = {".xlsx": "xlsx", ".xls": "xlsx", ".tsv": "tsv", ".csv": "csv"}


def _detect_format(path: str, explicit: Optional[str]) -> str:
    if explicit and explicit != "auto":
        return explicit
    suffix = Path(path).suffix.lower()
    fmt = _FORMAT_BY_SUFFIX.get(suffix)
    if fmt is None:
        raise ValueError(
            f"Cannot infer format from extension {suffix!r} for {path!r}; "
            "pass --input-format/--output-format explicitly."
        )
    return fmt


def _read_excel(path: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.active
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            try:
                idx = int(sheet) - 1
                ws = wb.worksheets[idx]
            except (ValueError, IndexError):
                raise ValueError(
                    f"Worksheet {sheet!r} not found in {path!r}. Available sheets: {wb.sheetnames}"
                )

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = [str(h).strip() if h is not None else "" for h in header]

        rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue
            rows.append({col: ("" if val is None else str(val)) for col, val in zip(fieldnames, raw_row)})
    finally:
        wb.close()

    return fieldnames, rows


def _read_delimited(path: str, sep: str) -> tuple[list[str], list[dict[str, str]]]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh, delimiter=sep)
        if reader.fieldnames is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = list(reader.fieldnames)
        rows = [dict(row) for row in reader]
    return fieldnames, rows


def _read_rows(path: str, fmt: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    if fmt == "xlsx":
        return _read_excel(path, sheet)
    return _read_delimited(path, "\t" if fmt == "tsv" else ",")


def _write_excel(path: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])
    wb.save(path)


def _write_delimited(path: str, sep: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, delimiter=sep, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_rows(path: str, fmt: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    if fmt == "xlsx":
        _write_excel(path, fieldnames, rows)
        return
    _write_delimited(path, "\t" if fmt == "tsv" else ",", fieldnames, rows)


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def fix_missing_ref_allele(
    input_file: str,
    output_file: str,
    *,
    input_format: Optional[str] = None,
    output_format: Optional[str] = None,
    sheet: Optional[str] = None,
    coding_col: str = "coding variant",
    seq_col: str = "seq",
    reference_col: str = "reference",
    on_unresolved: str = "error",
) -> dict:
    """Reconstruct missing reference letters in *input_file*, writing *output_file*.

    Returns a stats dict: ``checked`` (malformed-shaped cells found), ``fixed``.
    """
    in_fmt = _detect_format(input_file, input_format)
    out_fmt = _detect_format(output_file, output_format)

    fieldnames, rows = _read_rows(input_file, in_fmt, sheet)

    for col in (coding_col, seq_col, reference_col):
        if col not in fieldnames:
            raise ValueError(f"Column {col!r} not found in input; available columns: {fieldnames}")

    checked = 0
    fixed = 0
    for i, row in enumerate(rows, start=2):
        value = row[coding_col]
        if not value or ":" not in value:
            continue
        accession, rest = value.split(":", 1)
        match = _MISSING_REF_RE.match(rest)
        if not match:
            continue
        checked += 1

        seq = row[seq_col]
        reference = row[reference_col]
        offset = _mismatch_offset(seq, reference)
        alt = match.group("alt").upper()

        message = None
        if offset is None:
            message = (
                f"row {i}: {value!r} is missing its reference letter, but seq/reference "
                f"don't have exactly one mismatch (seq={seq!r}, reference={reference!r})"
            )
        elif seq[offset].upper() != alt:
            message = (
                f"row {i}: {value!r} claims alt {alt!r}, but the seq/reference mismatch "
                f"at offset {offset} is {seq[offset]!r} (reference={reference[offset]!r})"
            )

        if message:
            if on_unresolved == "error":
                raise ValueError(message)
            logger.warning(message)
            continue

        ref = reference[offset].upper()
        utr = match.group("utr") or ""
        pos_label = f"{utr}{match.group('base')}"
        if match.group("isign"):
            pos_label += f"{match.group('isign')}{match.group('offset')}"
        new_value = f"{accession}:c.{pos_label}{ref}>{alt}"
        logger.info("row %d: %s -> %s", i, value, new_value)
        row[coding_col] = new_value
        fixed += 1

    _write_rows(output_file, out_fmt, fieldnames, rows)
    stats = {"checked": checked, "fixed": fixed}
    logger.info(
        "Found %d malformed (missing-reference) cell(s); reconstructed %d; wrote %d rows to %s",
        checked, fixed, len(rows), output_file,
    )
    return stats


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Reconstruct a missing reference-allele letter in malformed coding-HGVS "
            "substitutions (e.g. 'c.865-4>G' -> 'c.865-4C>G'), using the row's own "
            "seq/reference columns."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("input", help="Input file (csv, tsv, or xlsx).")
    p.add_argument("output", help="Output file (csv, tsv, or xlsx).")

    p.add_argument("--input-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--output-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--sheet", default=None,
                    help="Worksheet name or 1-based index for xlsx input (default: active/first sheet).")

    p.add_argument("--coding-col", default="coding variant", metavar="COL")
    p.add_argument("--seq-col", default="seq", metavar="COL")
    p.add_argument("--reference-col", default="reference", metavar="COL")

    p.add_argument("--on-unresolved", choices=["error", "leave"], default="error",
                    help="'error' stops the run if a malformed cell's reference letter can't "
                         "be reconstructed (ambiguous or inconsistent seq/reference); 'leave' "
                         "logs a warning and leaves that cell unchanged.")

    p.add_argument("--csv-field-size-limit", type=int, default=131072)
    p.add_argument("--log-level", default="WARNING",
                    choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"])
    return p


def main(argv: Optional[list[str]] = None) -> None:
    parser = _build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        stream=sys.stderr,
    )

    csv.field_size_limit(args.csv_field_size_limit)

    stats = fix_missing_ref_allele(
        args.input,
        args.output,
        input_format=args.input_format,
        output_format=args.output_format,
        sheet=args.sheet,
        coding_col=args.coding_col,
        seq_col=args.seq_col,
        reference_col=args.reference_col,
        on_unresolved=args.on_unresolved,
    )

    print(
        f"{args.input}: found {stats['checked']} malformed (missing-reference) cell(s), "
        f"reconstructed {stats['fixed']}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
