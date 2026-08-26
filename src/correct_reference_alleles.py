"""Correct coding-HGVS substitutions whose reported reference allele doesn't match
the true RefSeq transcript sequence.

Some MAVE/SGE target constructs contain fixed, deliberately engineered edits (e.g.
silent recoding to remove a restriction site) that get folded into the assay's own
local target/reference sequence. When the systematic mutagenesis scan also mutates
one of those already-edited positions, the resulting variant call inherits the
*engineered* reference base rather than the transcript's true reference base. For
example, if a fixed edit changes the true C at c.462 to A in the target, the scan
can report "c.462A>C" (which actually just restores the true base) and "c.462A>T"
(a genuine variant), when the biologically correct descriptions are "c.462="
(HGVS "predicted no change" notation) and "c.462C>T" respectively.

This script looks up the true reference base at each single-nucleotide-substitution
position directly from the RefSeq transcript/genomic sequence (via UTA), and
rewrites the call:

  - True reference matches the reported reference: left unchanged (no fixed edit
    at this position).
  - True reference matches the reported alternate: rewritten to "ACCESSION:c.<pos>=".
  - Otherwise: rewritten to "ACCESSION:c.<pos><true_ref>><alt>", correcting the
    reference letter and keeping the alternate.

Three kinds of single-nucleotide-substitution positions are checked, each read from
the source appropriate to it:
  - coding  (``c.462``):        read from the transcript's spliced cDNA sequence.
  - UTR     (``c.-45``, ``c.*12``): read from the same cDNA sequence (UTRs are
                                     exonic too), offset from the CDS boundaries.
  - intronic (``c.359-4``, ``c.-45+3``, ``c.*12-5``): mapped to a genomic
    coordinate via ``hgvs``'s c-to-g mapper (which tolerates a mismatched input
    reference for intronic positions — it's only used for the position math
    here), then read from the genomic reference sequence.
Other coding-HGVS cells (ranges, "c.=" cells, indels, etc.) are passed through
unchanged.

Multi-variant haplotypes (``ACCESSION:c.[2523A>C;2533A>T]``, in the correctly
bracketed form — see ``translate_hgvs_accessions.py`` for normalizing the
malformed unbracketed form some source data uses) are unpacked and each
substitution component is checked/corrected independently. A component that
turns out to be a true no-change is dropped from the haplotype's variant list
(kept in its position-qualified "<pos>=" form only if every component
collapses this way); reduced to one remaining component, the brackets are
dropped. Non-substitution components (indels, etc.) within a haplotype are
left as-is.

Requires UTA_DB_URL in the environment (or --uta-db-url).
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# UTA-backed reference lookup
# ---------------------------------------------------------------------------

# Matches a single-nucleotide-substitution coding position in any of its forms:
#   c.462            coding
#   c.-45 / c.*12    5'/3' UTR
#   c.359-4          intronic (offset from a coding position)
#   c.-45+3 / c.*12-5  intronic (offset from a UTR position)
_POSITION_RE = re.compile(
    r"^c\.(?P<utr>[-*])?(?P<base>\d+)(?:(?P<isign>[-+])(?P<offset>\d+))?"
    r"(?P<ref>[ACGTNacgtn])>(?P<alt>[ACGTNacgtn])$"
)

# A correctly bracketed multi-variant haplotype: "c.[2523A>C;2533A>T]".
_HAPLOTYPE_RE = re.compile(r"^c\.\[(.+)\]$")

_COMPLEMENT = {"A": "T", "T": "A", "C": "G", "G": "C", "N": "N"}


def _complement_base(base: str) -> str:
    return _COMPLEMENT.get(base, base)


def _classify_position(utr: Optional[str], isign: Optional[str]) -> str:
    if isign:
        return "intronic"
    if utr:
        return "utr"
    return "coding"


class _UtaRefResolver:
    """Looks up true reference bases from UTA, caching per transcript.

    Two kinds of coding positions are supported:
      - plain CDS positions (``c.462``): read directly out of the transcript's
        own (spliced) cDNA sequence.
      - intronic offset positions (``c.359-4``, ``c.1017+1``): mapped to a
        genomic coordinate via ``hgvs``'s c-to-g mapper (which tolerates a
        mismatched input reference for intronic positions — it's only used for
        the position math here), then read out of the genomic reference
        sequence and complemented back to the transcript strand if needed.
    """

    def __init__(self, uta_db_url: Optional[str] = None) -> None:
        import hgvs.dataproviders.uta  # type: ignore[import-untyped]
        import hgvs.parser  # type: ignore[import-untyped]
        import hgvs.variantmapper  # type: ignore[import-untyped]

        db_url = (uta_db_url or os.environ.get("UTA_DB_URL") or "").strip()
        if not db_url:
            raise RuntimeError("UTA_DB_URL is required to look up true reference alleles.")
        self._hdp = hgvs.dataproviders.uta.connect(db_url)
        self._parser = hgvs.parser.Parser()
        self._vm = hgvs.variantmapper.VariantMapper(self._hdp)

        self._seq_cache: dict[str, str] = {}
        self._cds_start_cache: dict[str, int] = {}
        self._cds_end_cache: dict[str, int] = {}
        self._alignment_cache: dict[str, tuple[str, str]] = {}
        self._exon_cache: dict[str, list] = {}
        self._strand_cache: dict[str, int] = {}
        self._genomic_window_cache: dict[str, tuple[str, int]] = {}

    @staticmethod
    def _field(row, name: str):
        """UTA data-provider rows support both attribute and item access; try both."""
        val = getattr(row, name, None)
        if val is None and hasattr(row, "__getitem__"):
            try:
                val = row[name]
            except (KeyError, TypeError):
                val = None
        return val

    def _get_alignment(self, tx_ac: str) -> tuple[str, str]:
        """Return the (alt_ac, alt_aln_method) UTA uses for this transcript's alignment."""
        if tx_ac not in self._alignment_cache:
            mapping_options = self._hdp.get_tx_mapping_options(tx_ac)
            if not mapping_options:
                raise ValueError(f"No mapping options found in UTA for transcript {tx_ac!r}.")
            self._alignment_cache[tx_ac] = (mapping_options[0][1], mapping_options[0][2])
        return self._alignment_cache[tx_ac]

    def _get_full_seq(self, tx_ac: str) -> str:
        if tx_ac not in self._seq_cache:
            seq = self._hdp.get_seq(tx_ac)
            if not seq:
                raise ValueError(f"No sequence found in UTA for transcript {tx_ac!r}.")
            self._seq_cache[tx_ac] = seq.upper()
        return self._seq_cache[tx_ac]

    def _get_cds_start_i(self, tx_ac: str) -> int:
        """Return the 0-based interbase CDS start position in the transcript sequence.

        ``c.1`` is at transcript sequence index ``cds_start_i`` (0-based).
        """
        if tx_ac not in self._cds_start_cache:
            alt_ac, alt_aln_method = self._get_alignment(tx_ac)
            tx_info = self._hdp.get_tx_info(tx_ac, alt_ac, alt_aln_method)
            if tx_info is None:
                raise ValueError(f"No tx_info found in UTA for transcript {tx_ac!r}.")
            cds_start = self._field(tx_info, "cds_start_i")
            if cds_start is None:
                raise ValueError(f"No cds_start_i found in UTA for transcript {tx_ac!r}.")
            self._cds_start_cache[tx_ac] = int(cds_start)
        return self._cds_start_cache[tx_ac]

    def _get_cds_end_i(self, tx_ac: str) -> int:
        """Return the 0-based interbase CDS end position in the transcript sequence.

        ``c.*1`` (the first 3'UTR base) is at transcript sequence index ``cds_end_i``.
        """
        if tx_ac not in self._cds_end_cache:
            alt_ac, alt_aln_method = self._get_alignment(tx_ac)
            tx_info = self._hdp.get_tx_info(tx_ac, alt_ac, alt_aln_method)
            if tx_info is None:
                raise ValueError(f"No tx_info found in UTA for transcript {tx_ac!r}.")
            cds_end = self._field(tx_info, "cds_end_i")
            if cds_end is None:
                raise ValueError(f"No cds_end_i found in UTA for transcript {tx_ac!r}.")
            self._cds_end_cache[tx_ac] = int(cds_end)
        return self._cds_end_cache[tx_ac]

    def _get_exons(self, tx_ac: str) -> list:
        if tx_ac not in self._exon_cache:
            alt_ac, alt_aln_method = self._get_alignment(tx_ac)
            exons = self._hdp.get_tx_exons(tx_ac, alt_ac, alt_aln_method)
            if not exons:
                raise ValueError(f"No exon structure found in UTA for transcript {tx_ac!r}.")
            self._exon_cache[tx_ac] = exons
        return self._exon_cache[tx_ac]

    def _get_strand(self, tx_ac: str) -> int:
        if tx_ac not in self._strand_cache:
            strand = self._field(self._get_exons(tx_ac)[0], "alt_strand")
            if strand is None:
                raise ValueError(f"No strand found in UTA for transcript {tx_ac!r}.")
            self._strand_cache[tx_ac] = int(strand)
        return self._strand_cache[tx_ac]

    def _get_genomic_window(self, tx_ac: str) -> tuple[str, int]:
        """Return (sequence, window_start_0based) for the genomic span covering all of
        this transcript's exons (and therefore every intron between them) — fetched
        once per transcript and sliced locally, rather than one sequence-fetcher round
        trip per intronic position.
        """
        if tx_ac not in self._genomic_window_cache:
            alt_ac, _ = self._get_alignment(tx_ac)
            exons = self._get_exons(tx_ac)
            starts = [self._field(e, "alt_start_i") for e in exons]
            ends = [self._field(e, "alt_end_i") for e in exons]
            if any(v is None for v in starts + ends):
                raise ValueError(f"Incomplete exon coordinates in UTA for transcript {tx_ac!r}.")
            window_start = min(int(v) for v in starts)
            window_end = max(int(v) for v in ends)
            seq = self._hdp.get_seq(alt_ac, window_start, window_end)
            if not seq:
                raise ValueError(f"No genomic sequence found at {alt_ac}:{window_start}-{window_end}.")
            self._genomic_window_cache[tx_ac] = (seq.upper(), window_start)
        return self._genomic_window_cache[tx_ac]

    def true_exonic_base(self, tx_ac: str, utr: Optional[str], base: int) -> str:
        """Return the transcript's true nucleotide at an exonic position: coding
        (``utr=None``, 1-based CDS position), 5'UTR (``utr="-"``), or 3'UTR
        (``utr="*"``). Read directly out of the spliced cDNA sequence.
        """
        seq = self._get_full_seq(tx_ac)
        cds_start_i = self._get_cds_start_i(tx_ac)
        if utr == "-":
            idx = cds_start_i - base
        elif utr == "*":
            idx = self._get_cds_end_i(tx_ac) + base - 1
        else:
            idx = cds_start_i + base - 1
        if idx < 0 or idx >= len(seq):
            raise ValueError(
                f"Position c.{utr or ''}{base} is out of bounds for transcript {tx_ac!r}."
            )
        return seq[idx]

    def true_intronic_base(self, tx_ac: str, intron_pos: str, ref: str, alt: str) -> str:
        """Return the transcript-strand true nucleotide at intronic position *intron_pos*
        (e.g. ``"359-4"``, ``"-45+3"``, or ``"*12-5"``). *ref*/*alt* are only used to
        build a syntactically valid HGVS variant for position mapping — the mapper
        does not validate them for intronic positions, so a mismatched *ref* is fine."""
        alt_ac, _ = self._get_alignment(tx_ac)
        var_c = self._parser.parse_hgvs_variant(f"{tx_ac}:c.{intron_pos}{ref}>{alt}")
        var_g = self._vm.c_to_g(var_c, alt_ac)
        g_pos = var_g.posedit.pos.start.base

        window_seq, window_start = self._get_genomic_window(tx_ac)
        idx = g_pos - 1 - window_start
        if idx < 0 or idx >= len(window_seq):
            raise ValueError(
                f"Intronic position c.{intron_pos} (genomic {alt_ac}:{g_pos}) falls outside "
                f"the exon-spanning window fetched for transcript {tx_ac!r}."
            )
        genomic_base = window_seq[idx]
        strand = self._get_strand(tx_ac)
        return _complement_base(genomic_base) if strand == -1 else genomic_base


def _iter_position_matches(value: str):
    """Yield a ``_POSITION_RE`` match for each substitution component in *value*
    (handling both a single variant and a bracketed ``c.[...]`` haplotype)."""
    if not value or ":" not in value:
        return
    _, rest = value.split(":", 1)
    haplotype_match = _HAPLOTYPE_RE.match(rest)
    if haplotype_match:
        for component in haplotype_match.group(1).split(";"):
            m = _POSITION_RE.match(f"c.{component.strip()}")
            if m:
                yield m
        return
    m = _POSITION_RE.match(rest)
    if m:
        yield m


def _correct_component(
    accession: str,
    component: str,
    resolver: _UtaRefResolver,
    on_unresolved: str,
    row_context: str,
) -> tuple[str, bool, Optional[str], Optional[str]]:
    """Check/correct one HGVS coding-change component (no leading "c.", e.g. "462A>C").

    Returns (new_component, changed, position_key, category). *new_component* is
    always a valid standalone c.-suffix — never a bare, position-less "=". For a
    true no-change (true reference matches the reported alternate), it's
    ``"<pos>="`` (e.g. "462="), the HGVS "predicted no change" form, which keeps
    the position so distinct no-change positions don't collide with each other
    once rendered. Callers assembling a haplotype should drop components ending
    in "=" from the bracketed list (a no-change component isn't an actual
    co-occurring variant) but keep the position-qualified form if the whole
    haplotype collapses to no real changes. Components that aren't a recognized
    single-nucleotide substitution (indels, etc.) are returned unchanged with
    ``changed=False``.
    """
    match = _POSITION_RE.match(f"c.{component}")
    if not match:
        return component, False, None, None

    utr = match.group("utr") or ""
    base = int(match.group("base"))
    isign = match.group("isign")
    offset = match.group("offset")
    ref = match.group("ref").upper()
    alt = match.group("alt").upper()

    pos_label = f"{utr}{base}"
    if isign:
        pos_label += f"{isign}{offset}"
    category = _classify_position(utr, isign)

    try:
        if isign:
            true_ref = resolver.true_intronic_base(accession, pos_label, ref, alt)
        else:
            true_ref = resolver.true_exonic_base(accession, utr or None, base)
    except Exception as exc:
        message = (
            f"{row_context}: could not resolve true reference base for "
            f"{accession}:c.{component!r}: {exc}"
        )
        if on_unresolved == "error":
            raise ValueError(message) from exc
        logger.warning(message)
        return component, False, None, None

    if true_ref == ref:
        return component, False, None, None
    position_key = f"{accession}:c.{pos_label}"
    if true_ref == alt:
        return f"{pos_label}=", True, position_key, category
    return f"{pos_label}{true_ref}>{alt}", True, position_key, category


def _correct_value(
    value: str,
    resolver: _UtaRefResolver,
    on_unresolved: str,
    row_context: str,
) -> tuple[str, list[tuple[str, str]]]:
    """Return (possibly-rewritten value, [(position_key, category), ...] for each
    corrected component)."""
    if not value or ":" not in value:
        return value, []
    accession, rest = value.split(":", 1)

    haplotype_match = _HAPLOTYPE_RE.match(rest)
    if haplotype_match:
        components = [c.strip() for c in haplotype_match.group(1).split(";")]
        real_kept: list[str] = []
        nochange_kept: list[str] = []
        corrections: list[tuple[str, str]] = []
        for component in components:
            new_component, changed, position_key, category = _correct_component(
                accession, component, resolver, on_unresolved, row_context
            )
            if changed:
                corrections.append((position_key, category))
            if new_component.endswith("="):
                # A resolved-to-no-change component isn't an actual co-occurring
                # variant, so it doesn't belong in the haplotype's variant list —
                # unless *every* component collapses this way, in which case we
                # keep the position-qualified no-change forms rather than losing
                # all position information to a single ambiguous "c.=".
                nochange_kept.append(new_component)
            else:
                real_kept.append(new_component)
        if not corrections:
            return value, []
        kept = real_kept or nochange_kept
        if len(kept) == 1:
            new_value = f"{accession}:c.{kept[0]}"
        else:
            new_value = f"{accession}:c.[{';'.join(kept)}]"
        return new_value, corrections

    if not rest.startswith("c."):
        return value, []
    new_component, changed, position_key, category = _correct_component(
        accession, rest[2:], resolver, on_unresolved, row_context
    )
    if not changed:
        return value, []
    return f"{accession}:c.{new_component}", [(position_key, category)]


# ---------------------------------------------------------------------------
# File I/O (csv, tsv, xlsx) — same conventions as translate_hgvs_accessions.py
# ---------------------------------------------------------------------------

_FORMAT_BY_SUFFIX = {".xlsx": "xlsx", ".xls": "xlsx", ".tsv": "tsv", ".csv": "csv"}


def _detect_format(path: str, explicit: Optional[str]) -> str:
    if explicit and explicit != "auto":
        return explicit
    suffix = Path(path).suffix.lower()
    fmt = _FORMAT_BY_SUFFIX.get(suffix)
    if fmt is None:
        raise ValueError(
            f"Cannot infer format from extension {suffix!r} for {path!r}; "
            "pass --input-format/--output-format explicitly."
        )
    return fmt


def _read_excel(path: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.active
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            try:
                idx = int(sheet) - 1
                ws = wb.worksheets[idx]
            except (ValueError, IndexError):
                raise ValueError(
                    f"Worksheet {sheet!r} not found in {path!r}. Available sheets: {wb.sheetnames}"
                )

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = [str(h).strip() if h is not None else "" for h in header]

        rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue
            rows.append({col: ("" if val is None else str(val)) for col, val in zip(fieldnames, raw_row)})
    finally:
        wb.close()

    return fieldnames, rows


def _read_delimited(path: str, sep: str) -> tuple[list[str], list[dict[str, str]]]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh, delimiter=sep)
        if reader.fieldnames is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = list(reader.fieldnames)
        rows = [dict(row) for row in reader]
    return fieldnames, rows


def _read_rows(path: str, fmt: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    if fmt == "xlsx":
        return _read_excel(path, sheet)
    return _read_delimited(path, "\t" if fmt == "tsv" else ",")


def _write_excel(path: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])
    wb.save(path)


def _write_delimited(path: str, sep: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, delimiter=sep, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_rows(path: str, fmt: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    if fmt == "xlsx":
        _write_excel(path, fieldnames, rows)
        return
    _write_delimited(path, "\t" if fmt == "tsv" else ",", fieldnames, rows)


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

_CATEGORIES = ("coding", "intronic", "utr")


def correct_reference_alleles(
    input_file: str,
    output_file: str,
    *,
    input_format: Optional[str] = None,
    output_format: Optional[str] = None,
    sheet: Optional[str] = None,
    coding_col: str = "coding variant",
    uta_db_url: Optional[str] = None,
    on_unresolved: str = "error",
) -> dict:
    """Correct reference-allele mismatches in *input_file*, writing *output_file*.

    Returns a stats dict with overall ``checked``/``corrected_rows``/``corrected_positions``
    counts, a ``by_category`` breakdown (each of "coding", "intronic", "utr") with the
    same three counts, and ``corrections``: a list of ``{"original": ..., "corrected": ...}``
    dicts, one per row whose coding-variant cell changed (the whole cell value, before
    and after — so a haplotype's bracketed group counts as a single entry).
    """
    in_fmt = _detect_format(input_file, input_format)
    out_fmt = _detect_format(output_file, output_format)

    fieldnames, rows = _read_rows(input_file, in_fmt, sheet)

    if coding_col not in fieldnames:
        raise ValueError(f"Column {coding_col!r} not found in input; available columns: {fieldnames}")

    resolver = _UtaRefResolver(uta_db_url)

    checked = {cat: 0 for cat in _CATEGORIES}
    corrected_rows = {cat: 0 for cat in _CATEGORIES}
    corrected_positions: dict[str, set[str]] = {cat: set() for cat in _CATEGORIES}
    corrections_log: list[dict[str, str]] = []

    for i, row in enumerate(rows, start=2):
        value = row[coding_col]
        for m in _iter_position_matches(value):
            checked[_classify_position(m.group("utr"), m.group("isign"))] += 1
        new_value, corrections = _correct_value(value, resolver, on_unresolved, f"row {i}")
        if corrections:
            logger.info("%s: %s -> %s", f"row {i}", value, new_value)
            corrections_log.append({"original": value, "corrected": new_value})
            for position_key, category in corrections:
                corrected_rows[category] += 1
                corrected_positions[category].add(position_key)
        row[coding_col] = new_value

    _write_rows(output_file, out_fmt, fieldnames, rows)

    by_category = {
        cat: {
            "checked": checked[cat],
            "corrected_rows": corrected_rows[cat],
            "corrected_positions": len(corrected_positions[cat]),
        }
        for cat in _CATEGORIES
    }
    stats = {
        "checked": sum(checked.values()),
        "corrected_rows": sum(corrected_rows.values()),
        "corrected_positions": sum(len(s) for s in corrected_positions.values()),
        "by_category": by_category,
        "corrections": corrections_log,
    }
    logger.info(
        "Checked %d substitution row(s); corrected %d row(s) at %d position(s) with a "
        "fixed-edit reference mismatch (coding=%d/%d/%d, intronic=%d/%d/%d, utr=%d/%d/%d "
        "[checked/corrected/positions]); wrote %d rows to %s",
        stats["checked"], stats["corrected_rows"], stats["corrected_positions"],
        by_category["coding"]["checked"], by_category["coding"]["corrected_rows"], by_category["coding"]["corrected_positions"],
        by_category["intronic"]["checked"], by_category["intronic"]["corrected_rows"], by_category["intronic"]["corrected_positions"],
        by_category["utr"]["checked"], by_category["utr"]["corrected_rows"], by_category["utr"]["corrected_positions"],
        len(rows), output_file,
    )
    return stats


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Correct coding-HGVS substitutions whose reported reference allele doesn't "
            "match the true RefSeq transcript sequence (e.g. from fixed edits baked into "
            "a MAVE/SGE target), using UTA as the source of truth."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("input", help="Input file (csv, tsv, or xlsx).")
    p.add_argument("output", help="Output file (csv, tsv, or xlsx).")

    p.add_argument("--input-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--output-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--sheet", default=None,
                    help="Worksheet name or 1-based index for xlsx input (default: active/first sheet).")

    p.add_argument("--coding-col", default="coding variant", metavar="COL",
                    help="Column with coding HGVS (ACCESSION:c.change) to check/correct.")

    p.add_argument("--uta-db-url", default=None, metavar="URL",
                    help="UTA database URL. Defaults to the UTA_DB_URL environment variable.")
    p.add_argument("--on-unresolved", choices=["error", "leave"], default="error",
                    help="'error' stops the run if a substitution's true reference base can't "
                         "be resolved (e.g. transcript not in UTA); 'leave' logs a warning and "
                         "leaves that cell unchanged.")

    p.add_argument("--summary-file", default=None, metavar="FILE",
                    help="Write correction counts to FILE as shell-sourceable "
                         "'checked=N', 'corrected_rows=N', 'corrected_positions=N' lines, "
                         "in addition to the summary always printed to stderr.")
    p.add_argument("--corrections-file", default=None, metavar="FILE",
                    help="Write a two-column TSV (original, corrected) to FILE, one row per "
                         "coding-variant cell that was rewritten (a haplotype's bracketed "
                         "group counts as a single before/after row).")

    p.add_argument("--csv-field-size-limit", type=int, default=131072)
    p.add_argument("--log-level", default="WARNING",
                    choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"])
    return p


def main(argv: Optional[list[str]] = None) -> None:
    parser = _build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        stream=sys.stderr,
    )
    # The hgvs library logs an expected, harmless INFO message for every intronic
    # position it maps (it can't validate/update the ref for those); keep that out
    # of our own INFO-level output.
    logging.getLogger("hgvs").setLevel(logging.WARNING)

    csv.field_size_limit(args.csv_field_size_limit)

    stats = correct_reference_alleles(
        args.input,
        args.output,
        input_format=args.input_format,
        output_format=args.output_format,
        sheet=args.sheet,
        coding_col=args.coding_col,
        uta_db_url=args.uta_db_url,
        on_unresolved=args.on_unresolved,
    )

    print(
        f"{args.input}: corrected {stats['corrected_rows']} variant(s) at "
        f"{stats['corrected_positions']} position(s) (checked {stats['checked']} "
        "substitution row(s))",
        file=sys.stderr,
    )
    for cat in _CATEGORIES:
        c = stats["by_category"][cat]
        print(
            f"  {cat}: corrected {c['corrected_rows']} variant(s) at "
            f"{c['corrected_positions']} position(s) (checked {c['checked']} row(s))",
            file=sys.stderr,
        )

    if args.summary_file:
        with open(args.summary_file, "w", encoding="utf-8") as fh:
            fh.write(f"checked={stats['checked']}\n")
            fh.write(f"corrected_rows={stats['corrected_rows']}\n")
            fh.write(f"corrected_positions={stats['corrected_positions']}\n")
            for cat in _CATEGORIES:
                c = stats["by_category"][cat]
                fh.write(f"checked_{cat}={c['checked']}\n")
                fh.write(f"corrected_rows_{cat}={c['corrected_rows']}\n")
                fh.write(f"corrected_positions_{cat}={c['corrected_positions']}\n")

    if args.corrections_file:
        with open(args.corrections_file, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=["original", "corrected"], delimiter="\t")
            writer.writeheader()
            writer.writerows(stats["corrections"])


if __name__ == "__main__":
    main()
