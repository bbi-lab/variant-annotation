"""Expand bare "c.=" (unchanged) coding-HGVS cells into a position or range.

MAVE wild-type/"unchanged" control rows report a coding-variant cell of
``ACCESSION:c.=`` with no position — even though distinct control rows
represent different local windows of the target (their seq/reference columns
differ). This script infers each control row's coding-position range from its
own reference window text:

1. Every OTHER row in the table whose coding-variant cell is a resolvable
   single-nucleotide substitution (e.g. ``c.208G>A``) pins down the absolute
   coding position of every base in *its* reference window (position minus the
   offset of the substituted base within that window = the window's start).
2. For an unchanged row, if some other row's reference window is textually
   identical to this row's reference, that window's inferred start applies
   directly.
3. Failing that, if this row's reference window is a substring of some other
   (longer) resolved window, the start is taken from that window's start plus
   the substring's offset within it.
4. The row's own reference length then gives the window's end.

The result is written as ``ACCESSION:c.<start>_<end>=`` (or
``ACCESSION:c.<start>=`` if the window is a single nucleotide). Rows whose
coding-variant cell is not exactly ``ACCESSION:c.=`` are left untouched.

Rows where no anchor can be found are, by default, treated as an error (see
--on-unresolved) rather than silently emitting a placeholder coordinate.
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Position inference
# ---------------------------------------------------------------------------

_SUBSTITUTION_RE = re.compile(r"^c\.(\d+)[ACGTNacgtn]>[ACGTNacgtn]$")


def _mismatch_offset(seq: str, ref: str) -> Optional[int]:
    """Return the single 0-based index where *seq* differs from *ref*, or None."""
    if len(seq) != len(ref):
        return None
    offsets = [i for i, (a, b) in enumerate(zip(seq, ref)) if a != b]
    return offsets[0] if len(offsets) == 1 else None


def _build_window_starts(
    rows: list[dict[str, str]],
    coding_col: str,
    seq_col: str,
    reference_col: str,
) -> dict[str, int]:
    """Map reference-window text -> inferred absolute coding start position.

    Built from rows whose coding-variant cell is a resolvable single-nucleotide
    substitution. Windows with mutually inconsistent inferred starts (from
    different substitution rows sharing identical reference text) are dropped
    with a warning rather than used.
    """
    candidates: dict[str, set[int]] = {}
    for row in rows:
        split = row[coding_col].split(":", 1)
        if len(split) != 2:
            continue
        match = _SUBSTITUTION_RE.match(split[1])
        if not match:
            continue
        ref_seq = row[reference_col]
        offset = _mismatch_offset(row[seq_col], ref_seq)
        if offset is None:
            continue
        pos = int(match.group(1))
        candidates.setdefault(ref_seq, set()).add(pos - offset)

    resolved: dict[str, int] = {}
    for ref_seq, starts in candidates.items():
        if len(starts) == 1:
            resolved[ref_seq] = next(iter(starts))
        else:
            logger.warning(
                "Reference window %r has inconsistent inferred start positions %s "
                "across substitution rows; not using it as an anchor.",
                ref_seq, sorted(starts),
            )
    return resolved


def _find_window_start(ref_seq: str, window_starts: dict[str, int]) -> Optional[int]:
    if ref_seq in window_starts:
        return window_starts[ref_seq]
    for other_ref, start in window_starts.items():
        idx = other_ref.find(ref_seq)
        if idx != -1:
            return start + idx
    return None


# ---------------------------------------------------------------------------
# File I/O (csv, tsv, xlsx) — same conventions as translate_hgvs_accessions.py
# ---------------------------------------------------------------------------

_FORMAT_BY_SUFFIX = {".xlsx": "xlsx", ".xls": "xlsx", ".tsv": "tsv", ".csv": "csv"}


def _detect_format(path: str, explicit: Optional[str]) -> str:
    if explicit and explicit != "auto":
        return explicit
    suffix = Path(path).suffix.lower()
    fmt = _FORMAT_BY_SUFFIX.get(suffix)
    if fmt is None:
        raise ValueError(
            f"Cannot infer format from extension {suffix!r} for {path!r}; "
            "pass --input-format/--output-format explicitly."
        )
    return fmt


def _read_excel(path: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.active
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            try:
                idx = int(sheet) - 1
                ws = wb.worksheets[idx]
            except (ValueError, IndexError):
                raise ValueError(
                    f"Worksheet {sheet!r} not found in {path!r}. Available sheets: {wb.sheetnames}"
                )

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = [str(h).strip() if h is not None else "" for h in header]

        rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue
            rows.append({col: ("" if val is None else str(val)) for col, val in zip(fieldnames, raw_row)})
    finally:
        wb.close()

    return fieldnames, rows


def _read_delimited(path: str, sep: str) -> tuple[list[str], list[dict[str, str]]]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh, delimiter=sep)
        if reader.fieldnames is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = list(reader.fieldnames)
        rows = [dict(row) for row in reader]
    return fieldnames, rows


def _read_rows(path: str, fmt: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    if fmt == "xlsx":
        return _read_excel(path, sheet)
    return _read_delimited(path, "\t" if fmt == "tsv" else ",")


def _write_excel(path: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])
    wb.save(path)


def _write_delimited(path: str, sep: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, delimiter=sep, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_rows(path: str, fmt: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    if fmt == "xlsx":
        _write_excel(path, fieldnames, rows)
        return
    _write_delimited(path, "\t" if fmt == "tsv" else ",", fieldnames, rows)


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def expand_unchanged_hgvs_c(
    input_file: str,
    output_file: str,
    *,
    input_format: Optional[str] = None,
    output_format: Optional[str] = None,
    sheet: Optional[str] = None,
    coding_col: str = "coding variant",
    seq_col: str = "seq",
    reference_col: str = "reference",
    on_unresolved: str = "error",
) -> None:
    in_fmt = _detect_format(input_file, input_format)
    out_fmt = _detect_format(output_file, output_format)

    fieldnames, rows = _read_rows(input_file, in_fmt, sheet)

    for col in (coding_col, seq_col, reference_col):
        if col not in fieldnames:
            raise ValueError(f"Column {col!r} not found in input; available columns: {fieldnames}")

    window_starts = _build_window_starts(rows, coding_col, seq_col, reference_col)

    n_expanded = 0
    unresolved: list[str] = []
    for i, row in enumerate(rows, start=2):
        value = row[coding_col]
        split = value.split(":", 1) if value else []
        if len(split) != 2 or split[1] != "c.=":
            continue
        accession = split[0]
        ref_seq = row[reference_col]
        start = _find_window_start(ref_seq, window_starts)
        if start is None:
            unresolved.append(f"row {i} ({accession}, reference={ref_seq!r})")
            continue
        end = start + len(ref_seq) - 1
        row[coding_col] = f"{accession}:c.{start}=" if start == end else f"{accession}:c.{start}_{end}="
        n_expanded += 1

    if unresolved:
        message = (
            f"{len(unresolved)} unchanged-variant row(s) could not be anchored to a position "
            f"(no identical or substring-matching reference window with a resolvable substitution "
            f"was found):\n  " + "\n  ".join(unresolved)
        )
        if on_unresolved == "error":
            raise ValueError(message)
        logger.warning("%s\nLeaving these cells as 'c.=' unchanged.", message)

    _write_rows(output_file, out_fmt, fieldnames, rows)
    logger.info("Expanded %d unchanged-variant row(s); wrote %d rows to %s", n_expanded, len(rows), output_file)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Expand bare 'ACCESSION:c.=' (unchanged/wild-type control) cells into "
            "'ACCESSION:c.<start>_<end>=', inferring the position from other rows' "
            "resolvable substitutions sharing the same reference window."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("input", help="Input file (csv, tsv, or xlsx).")
    p.add_argument("output", help="Output file (csv, tsv, or xlsx).")

    p.add_argument("--input-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--output-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--sheet", default=None,
                    help="Worksheet name or 1-based index for xlsx input (default: active/first sheet).")

    p.add_argument("--coding-col", default="coding variant", metavar="COL",
                    help="Column with coding HGVS (ACCESSION:c.change), including 'c.=' rows.")
    p.add_argument("--seq-col", default="seq", metavar="COL",
                    help="Column with the observed local read/window sequence.")
    p.add_argument("--reference-col", default="reference", metavar="COL",
                    help="Column with the local reference window sequence.")

    p.add_argument("--on-unresolved", choices=["error", "leave"], default="error",
                    help="'error' stops the run if a 'c.=' row can't be anchored to a position; "
                         "'leave' logs a warning and leaves it as 'c.=' unchanged.")

    p.add_argument("--csv-field-size-limit", type=int, default=131072)
    p.add_argument("--log-level", default="WARNING",
                    choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"])
    return p


def main(argv: Optional[list[str]] = None) -> None:
    parser = _build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        stream=sys.stderr,
    )

    csv.field_size_limit(args.csv_field_size_limit)

    expand_unchanged_hgvs_c(
        args.input,
        args.output,
        input_format=args.input_format,
        output_format=args.output_format,
        sheet=args.sheet,
        coding_col=args.coding_col,
        seq_col=args.seq_col,
        reference_col=args.reference_col,
        on_unresolved=args.on_unresolved,
    )


if __name__ == "__main__":
    main()
