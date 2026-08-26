"""Deduplicate MaveDB variant rows by coding-HGVS value.

MaveDB requires exactly one row per variant. Reference-allele correction can
introduce duplicate coding-variant values in (at least) two ways:

  - A multi-variant haplotype whose "real" component(s) get dropped during
    correction collapses to a single-variant call that coincides with an
    already-existing plain row for that same variant (e.g.
    "c.[2523A>C;2533A>T]" -> "c.2533A>T", when "c.2533A>T" already exists
    elsewhere in the file).
  - Two different positions both turn out to be pure "restores the true
    reference" calls and both collapse to the bare "c.=" no-change form,
    which carries no position information and so the two are
    indistinguishable and collide.

Duplicates that predate correction entirely (e.g. two overlapping tiled
windows both covering the same position, where one window's local reference
disagrees with the true transcript at some positions but not others) are
handled the same way, just possibly without a haplotype-derived row to prefer
dropping.

For each group of rows sharing the same coding-HGVS value, this script ranks
rows into tiers and keeps the best-ranked one (ties broken by file order),
writing the rest to a separate "omitted" file:
  1. Rows that were NOT originally a multi-variant haplotype (a bracketed
     "c.[...]" call) AND did not need reference-allele correction (their
     coding-HGVS value is unchanged from *original_file*) — i.e. rows whose
     reported reference already matched the true transcript/genomic sequence.
  2. Rows that were not originally a haplotype (regardless of whether they
     needed correction).
  3. Any row (all remaining candidates were haplotype-derived).

"Originally a haplotype" and "needed correction" are both determined from
*original_file*, a row-order-aligned file holding the coding-HGVS values
before correction (e.g. the pipeline's pre-correction intermediate) — it must
have the same number of rows, in the same order, as the input file.
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# File I/O (csv, tsv, xlsx) — same conventions as translate_hgvs_accessions.py
# ---------------------------------------------------------------------------

_FORMAT_BY_SUFFIX = {".xlsx": "xlsx", ".xls": "xlsx", ".tsv": "tsv", ".csv": "csv"}


def _detect_format(path: str, explicit: Optional[str]) -> str:
    if explicit and explicit != "auto":
        return explicit
    suffix = Path(path).suffix.lower()
    fmt = _FORMAT_BY_SUFFIX.get(suffix)
    if fmt is None:
        raise ValueError(
            f"Cannot infer format from extension {suffix!r} for {path!r}; "
            "pass the relevant --*-format option explicitly."
        )
    return fmt


def _read_excel(path: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.active
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            try:
                idx = int(sheet) - 1
                ws = wb.worksheets[idx]
            except (ValueError, IndexError):
                raise ValueError(
                    f"Worksheet {sheet!r} not found in {path!r}. Available sheets: {wb.sheetnames}"
                )

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = [str(h).strip() if h is not None else "" for h in header]

        rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue
            rows.append({col: ("" if val is None else str(val)) for col, val in zip(fieldnames, raw_row)})
    finally:
        wb.close()

    return fieldnames, rows


def _read_delimited(path: str, sep: str) -> tuple[list[str], list[dict[str, str]]]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh, delimiter=sep)
        if reader.fieldnames is None:
            raise ValueError(f"Input file {path!r} appears to be empty.")
        fieldnames = list(reader.fieldnames)
        rows = [dict(row) for row in reader]
    return fieldnames, rows


def _read_rows(path: str, fmt: str, sheet: Optional[str]) -> tuple[list[str], list[dict[str, str]]]:
    if fmt == "xlsx":
        return _read_excel(path, sheet)
    return _read_delimited(path, "\t" if fmt == "tsv" else ",")


def _write_excel(path: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])
    wb.save(path)


def _write_delimited(path: str, sep: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, delimiter=sep, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_rows(path: str, fmt: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    if fmt == "xlsx":
        _write_excel(path, fieldnames, rows)
        return
    _write_delimited(path, "\t" if fmt == "tsv" else ",", fieldnames, rows)


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def dedupe_mavedb_variants(
    input_file: str,
    original_file: str,
    output_file: str,
    omitted_file: str,
    *,
    coding_col: str = "coding variant",
    input_format: Optional[str] = None,
    original_format: Optional[str] = None,
    output_format: Optional[str] = None,
    omitted_format: Optional[str] = None,
    sheet: Optional[str] = None,
) -> dict:
    """Deduplicate *input_file* by *coding_col*, writing *output_file* and *omitted_file*.

    Returns a stats dict: ``input_rows``, ``output_rows``, ``omitted_rows``,
    ``duplicate_groups``.
    """
    in_fmt = _detect_format(input_file, input_format)
    orig_fmt = _detect_format(original_file, original_format)
    out_fmt = _detect_format(output_file, output_format)
    omit_fmt = _detect_format(omitted_file, omitted_format)

    fieldnames, rows = _read_rows(input_file, in_fmt, sheet)
    orig_fieldnames, orig_rows = _read_rows(original_file, orig_fmt, sheet)

    if coding_col not in fieldnames:
        raise ValueError(f"Column {coding_col!r} not found in {input_file!r}; available columns: {fieldnames}")
    if coding_col not in orig_fieldnames:
        raise ValueError(f"Column {coding_col!r} not found in {original_file!r}; available columns: {orig_fieldnames}")
    if len(rows) != len(orig_rows):
        raise ValueError(
            f"Row count mismatch: {input_file!r} has {len(rows)} rows but "
            f"{original_file!r} has {len(orig_rows)}; they must be row-aligned."
        )

    was_haplotype = ["[" in (orig_rows[i][coding_col] or "") for i in range(len(rows))]
    was_corrected = [orig_rows[i][coding_col] != rows[i][coding_col] for i in range(len(rows))]

    def _tier(i: int) -> int:
        if not was_haplotype[i] and not was_corrected[i]:
            return 0
        if not was_haplotype[i]:
            return 1
        return 2

    groups: dict[str, list[int]] = {}
    for i, row in enumerate(rows):
        groups.setdefault(row[coding_col], []).append(i)

    keep_indices: set[int] = set()
    omit_indices: list[int] = []
    duplicate_groups = 0
    for value, indices in groups.items():
        if len(indices) == 1:
            keep_indices.add(indices[0])
            continue
        duplicate_groups += 1
        keep = min(indices, key=lambda i: (_tier(i), i))
        keep_indices.add(keep)
        dropped = [i for i in indices if i != keep]
        omit_indices.extend(dropped)
        logger.info(
            "Duplicate %r: keeping row %d (tier %d), omitting row(s) %s",
            value, keep + 2, _tier(keep), [i + 2 for i in dropped],
        )

    kept_rows = [rows[i] for i in range(len(rows)) if i in keep_indices]
    omitted_rows = [rows[i] for i in sorted(omit_indices)]

    _write_rows(output_file, out_fmt, fieldnames, kept_rows)
    _write_rows(omitted_file, omit_fmt, fieldnames, omitted_rows)

    stats = {
        "input_rows": len(rows),
        "output_rows": len(kept_rows),
        "omitted_rows": len(omitted_rows),
        "duplicate_groups": duplicate_groups,
    }
    logger.info(
        "%d row(s) in, %d duplicate group(s), %d row(s) omitted, %d row(s) written to %s "
        "(omitted rows written to %s)",
        stats["input_rows"], stats["duplicate_groups"], stats["omitted_rows"],
        stats["output_rows"], output_file, omitted_file,
    )
    return stats


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Deduplicate MaveDB variant rows by coding-HGVS value, preferring to drop "
            "rows that were originally a multi-variant haplotype (see original_file)."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("input", help="MaveDB variant file (csv, tsv, or xlsx) to deduplicate.")
    p.add_argument("original", help="Row-aligned pre-correction file (same row count/order) "
                                     "used to detect which rows were originally haplotypes.")
    p.add_argument("output", help="Deduplicated output file.")
    p.add_argument("omitted", help="File to write the omitted (dropped duplicate) rows to.")

    p.add_argument("--coding-col", default="coding variant", metavar="COL")

    p.add_argument("--input-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--original-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--output-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--omitted-format", choices=["auto", "csv", "tsv", "xlsx"], default="auto")
    p.add_argument("--sheet", default=None,
                    help="Worksheet name or 1-based index for xlsx input (default: active/first sheet).")

    p.add_argument("--csv-field-size-limit", type=int, default=131072)
    p.add_argument("--log-level", default="WARNING",
                    choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"])
    return p


def main(argv: Optional[list[str]] = None) -> None:
    parser = _build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        stream=sys.stderr,
    )

    csv.field_size_limit(args.csv_field_size_limit)

    stats = dedupe_mavedb_variants(
        args.input,
        args.original,
        args.output,
        args.omitted,
        coding_col=args.coding_col,
        input_format=args.input_format,
        original_format=args.original_format,
        output_format=args.output_format,
        omitted_format=args.omitted_format,
        sheet=args.sheet,
    )

    print(
        f"{args.input}: {stats['duplicate_groups']} duplicate group(s), "
        f"{stats['omitted_rows']} row(s) omitted, {stats['output_rows']} row(s) kept "
        f"(from {stats['input_rows']} input rows)",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
